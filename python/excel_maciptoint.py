import openpyxl

def mac_to_int(mac_address):
    if isinstance(mac_address, int):
        return mac_address
    if mac_address:
        return int(mac_address.replace(':', ''), 16)
    return None

def ip_to_int(ip_address):
    if isinstance(ip_address, int):
        return ip_address
    if ip_address:
        segments = ip_address.split('.')
        return sum(int(segment) * (256 ** (3 - index)) for index, segment in enumerate(segments))
    return None


# xlsx 파일에서 MAC 및 IP 주소를 읽고, 정수로 변환한 후 업데이트하는 함수
def update_mac_ip_in_xlsx(file_path):
    # xlsx 파일 로드
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    # 데이터 행을 반복하여 처리
    for row in range(2, sheet.max_row + 1):  # 첫 번째 행은 헤더이므로 두 번째 행부터 시작
        # MAC 주소 변환
        mac_cell = sheet.cell(row=row, column=4)
        mac_int = mac_to_int(mac_cell.value)
        mac_cell.value = mac_int

        # IP 주소 변환
        ip_cell = sheet.cell(row=row, column=5)
        ip_int = ip_to_int(ip_cell.value)
        ip_cell.value = ip_int

    # 변경된 xlsx 파일 저장
    workbook.save(file_path)

update_mac_ip_in_xlsx('EMPLOYEE_LIST.xlsx')
