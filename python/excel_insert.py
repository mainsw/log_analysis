import openpyxl
import mysql.connector

# MySQL 연결 설정
config = {
    'user': 'root',
    'password': 'root',
    'host': 'localhost',
    'database': 'db',
}

def get_cell_value(cell):
    if cell.value is None:
        return None
    if isinstance(cell.value, str):
        return cell.value.strip()
    return cell.value

# xlsx 파일에서 데이터를 읽고, MySQL에 저장하는 함수
def save_data_to_mysql(xlsx_file_path):
    connection = mysql.connector.connect(**config)
    cursor = connection.cursor()
    
    # xlsx 파일 로드
    workbook = openpyxl.load_workbook(xlsx_file_path)
    sheet = workbook.active

    # 데이터 행을 반복하여 처리
    for row in range(2, sheet.max_row + 1):  # 첫 번째 행은 헤더이므로 두 번째 행부터 시작
        data = [
            get_cell_value(sheet.cell(row=row, column=1)),  # NUM
            get_cell_value(sheet.cell(row=row, column=2)),  # NAME
            get_cell_value(sheet.cell(row=row, column=3)),  # ACCOUNT
            get_cell_value(sheet.cell(row=row, column=4)),  # MAC
            get_cell_value(sheet.cell(row=row, column=5)),  # IP
            get_cell_value(sheet.cell(row=row, column=6))   # REGION
        ]
        cursor.execute('INSERT INTO employee (NUM, NAME, ACCOUNT, MAC, IP, REGION) VALUES (%s, %s, %s, %s, %s, %s)', data)
    
    connection.commit()
    cursor.close()
    connection.close()


save_data_to_mysql('EMPLOYEE_LIST.xlsx')
